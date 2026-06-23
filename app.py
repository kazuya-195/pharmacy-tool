import streamlit as st
import pandas as pd
import re
import time
import os
import subprocess
from io import BytesIO
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

# ============================================================
# 設定
# ============================================================
SEARCH_API   = "https://www.iryou.teikyouseido.mhlw.go.jp/znk-web/juminkanja/S2300/yakkyokuSearch"
RESULTS_BASE = "https://www.iryou.teikyouseido.mhlw.go.jp/znk-web/juminkanja/S2400/initialize?id="
PAGE_DELAY   = 2.5
DETAIL_DELAY = 1.5

REGION_MAP = {
    "北海道":"01_北海道",
    "青森県":"02_東北","岩手県":"02_東北","宮城県":"02_東北",
    "秋田県":"02_東北","山形県":"02_東北","福島県":"02_東北",
    "茨城県":"03_関東","栃木県":"03_関東","群馬県":"03_関東",
    "埼玉県":"03_関東","千葉県":"03_関東","東京都":"03_関東","神奈川県":"03_関東",
    "新潟県":"04_中部","富山県":"04_中部","石川県":"04_中部",
    "福井県":"04_中部","山梨県":"04_中部","長野県":"04_中部",
    "岐阜県":"04_中部","静岡県":"04_中部","愛知県":"04_中部",
    "三重県":"05_近畿","滋賀県":"05_近畿","京都府":"05_近畿",
    "大阪府":"05_近畿","兵庫県":"05_近畿","奈良県":"05_近畿","和歌山県":"05_近畿",
    "鳥取県":"06_中国四国","島根県":"06_中国四国","岡山県":"06_中国四国",
    "広島県":"06_中国四国","山口県":"06_中国四国",
    "徳島県":"06_中国四国","香川県":"06_中国四国","愛媛県":"06_中国四国","高知県":"06_中国四国",
    "福岡県":"07_九州沖縄","佐賀県":"07_九州沖縄","長崎県":"07_九州沖縄",
    "熊本県":"07_九州沖縄","大分県":"07_九州沖縄","宮崎県":"07_九州沖縄",
    "鹿児島県":"07_九州沖縄","沖縄県":"07_九州沖縄",
}

# ============================================================
# Seleniumドライバー
# ============================================================
def _find_bin(candidates):
    for name in candidates:
        try:
            path = subprocess.check_output(["which", name], text=True).strip()
            if path:
                return path
        except Exception:
            pass
    return None

def get_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1280,800")
    chromium = _find_bin(["chromium", "chromium-browser", "google-chrome-stable"])
    if chromium:
        options.binary_location = chromium
    chromedriver = _find_bin(["chromedriver", "chromium-chromedriver"])
    service = Service(chromedriver) if chromedriver else Service()
    return webdriver.Chrome(service=service, options=options)

# ============================================================
# Step1: Seleniumセッションを使ってAPIで検索ID取得
# ============================================================
def get_session_id(driver, keyword: str, debug: bool) -> str | None:
    """
    ★根本修正：requestsライブラリを廃止し、Seleniumブラウザ内でfetch APIを使う。
    
    なぜ：ナビィはsame-originリクエスト（ブラウザから直接）しか受け付けず、
    　　　requestsでクッキーをコピーしても HTTP 503 を返してしまう。
    　　　Selenium経由でJavaScript fetchを実行すれば、
    　　　ブラウザのセッション・クッキーをそのまま使えるため503を回避できる。
    """
    import urllib.parse

    base = "https://www.iryou.teikyouseido.mhlw.go.jp"
    driver.get(f"{base}/znk-web/juminkanja/S2300/initialize")
    time.sleep(4)

    if debug:
        st.image(driver.get_screenshot_as_png(), caption="① トップページ（セッション確立）")

    # クエリパラメータを組み立てる
    params = {
        "XCHARSET": "utf-8",
        "XPARAM": "keyword",
        "iyakuKbn": "2",
        "lang": "ja",
        "keywordType": "2",
        "keyword": keyword,
    }
    api_url = SEARCH_API + "?" + urllib.parse.urlencode(params)

    if debug:
        st.info(f"API URL: {api_url}")

    # ★ Seleniumブラウザ内でfetch → クッキーはブラウザのものをそのまま使う
    js_fetch = """
        var callback = arguments[0];
        var url = arguments[1];
        fetch(url, {
            method: 'GET',
            credentials: 'include',
            headers: {'Accept': 'application/json, text/plain, */*'}
        })
        .then(function(r) {
            return r.text().then(function(body) {
                return {status: r.status, body: body};
            });
        })
        .then(function(result) { callback({ok: true, result: result}); })
        .catch(function(err) { callback({ok: false, error: err.toString()}); });
    """

    for attempt in range(1, 4):
        try:
            if debug and attempt > 1:
                st.warning(f"APIリトライ {attempt}回目...")

            # execute_async_script でPromiseの結果を受け取る
            ret = driver.execute_async_script(js_fetch, api_url)

            if not ret or not ret.get("ok"):
                raise RuntimeError(f"fetchエラー: {ret.get('error', '不明')}")

            http_status = ret["result"]["status"]
            body        = ret["result"]["body"]

            if debug:
                st.info(f"APIレスポンス [HTTP {http_status}]: {body[:300]}")

            if http_status != 200:
                raise RuntimeError(f"HTTP {http_status} が返りました。bodyの先頭: {body[:100]}")

            import json
            data = json.loads(body)
            code = data.get("code")
            result_id = data.get("result", {}).get("id")

            if code != "0" or not result_id:
                st.error(f"APIエラー: code={code}, response={data}")
                return None

            return result_id

        except Exception as e:
            if debug:
                st.error(f"API呼び出し失敗 [{type(e).__name__}] (試行{attempt}/3): {e}")
            if attempt < 3:
                time.sleep(3 * attempt)

    st.error("API呼び出し3回とも失敗しました。デバッグモードでエラー詳細を確認してください。")
    return None

# ============================================================
# ページネーション補助関数
# ============================================================

def _get_current_s2430_urls(driver) -> set:
    """現在ページに表示中のS2430リンクURLセットを返す"""
    links = driver.find_elements(By.XPATH, "//a[contains(@href, 'S2430')]")
    return {l.get_attribute("href") for l in links if l.is_displayed() and l.get_attribute("href")}


def _find_next_button(driver, debug: bool):
    """
    >> または 次へ ボタンを返す。
    ★修正②：LI要素を除外し、A/button要素のみ対象にする。
    複数ある場合は最後のもの（= ページ下部のナビ）を選ぶ。
    """
    for xpath in [
        # A要素またはbutton要素の >> （LIは除外）
        "(//a | //button)[normalize-space(.)='>>' and not(@disabled)]",
        "(//a | //button)[normalize-space(.)='次へ' and not(@disabled)]",
        # span等で >> を囲んでいる場合の親A
        "//a[.//*[normalize-space(text())='>>'] and not(@disabled)]",
    ]:
        found = [
            b for b in driver.find_elements(By.XPATH, xpath)
            if b.is_displayed() and b.is_enabled()
        ]
        if found:
            if debug:
                tags = [(b.tag_name, b.text.strip()) for b in found]
                st.write(f"「>>」ボタン検出: {tags} → 最後の要素をクリック")
            return found[-1]
    return None


def _wait_for_page_change(driver, old_urls: set, timeout: int = 12) -> bool:
    """
    ★修正③：固定sleepではなくS2430リンクのURLセット変化でページ更新を検知する。
    最大timeout秒待機。変化したらTrue、タイムアウトでFalse。
    """
    for _ in range(timeout * 2):  # 0.5秒ずつポーリング
        time.sleep(0.5)
        new_urls = _get_current_s2430_urls(driver)
        # URLセットが変わった、かつ新しいURLが存在する場合に更新と判断
        if new_urls and new_urls != old_urls:
            return True
    return False

# ============================================================
# Step2: 結果ページから店舗リンクを収集
# ============================================================
def collect_urls(driver, session_id: str, status_text, debug: bool) -> list:
    driver.get(RESULTS_BASE + session_id)
    time.sleep(PAGE_DELAY)

    if debug:
        st.image(driver.get_screenshot_as_png(), caption="② 結果ページ（初期表示）")

    # モーダルを閉じる
    for sel in ["button.modalClose", "[class*='modal'] button", "button"]:
        try:
            for b in driver.find_elements(By.CSS_SELECTOR, sel):
                if b.is_displayed() and b.text.strip() in ("閉じる", "OK", "×"):
                    driver.execute_script("arguments[0].click();", b)
                    time.sleep(1)
                    break
        except Exception:
            pass

    time.sleep(1)

    all_urls    = []
    seen_hrefs  = set()
    page_num    = 1
    MAX_PAGES   = 20

    SKIP = {"ホーム","トップ","次へ","前へ","閉じる","戻る","条件を絞り込む",
            "全国の薬局","検索条件","お気に入り","ログイン","利用規約","関係者",
            "医療機関を探す","薬局を探す","キーワードで探す","アイコンの説明"}
    SKIP_PATTERNS = ["window","オブジェクト","//","function","undefined",
                     "null","Copyright","javascript","Script"]

    while page_num <= MAX_PAGES:
        # ---- 現在ページのS2430リンクを収集 ----
        js_links = driver.execute_script("""
            return Array.from(document.querySelectorAll('a')).map(l => ({
                text: (l.innerText || l.textContent || '').trim(),
                href: l.href || ''
            }));
        """)

        page_urls = []
        page_seen = set()
        for item in js_links:
            name = re.sub(r"\s+", " ", item.get("text", "")).strip()
            href = item.get("href", "")
            if (name and href
                    and len(name) >= 3
                    and "javascript" not in href
                    and href not in page_seen
                    and name not in SKIP
                    and not any(p in name for p in SKIP_PATTERNS)
                    and "iryou.teikyouseido" in href
                    and "S2430" in href):
                page_urls.append((name, href))
                page_seen.add(href)

        new_count = 0
        for name, href in page_urls:
            if href not in seen_hrefs:
                all_urls.append((name, href))
                seen_hrefs.add(href)
                new_count += 1

        status_text.text(f"収集中... ページ{page_num}: {new_count}件 / 累計{len(all_urls)}件")

        if debug:
            st.write(f"**ページ{page_num}**: 検出{len(page_urls)}件 / 新規{new_count}件 / 累計{len(all_urls)}件")
            if page_num == 1:
                st.image(driver.get_screenshot_as_png(), caption=f"ページ{page_num} 表示内容")

        # ---- 次ページへの遷移 ----
        # ★修正③：クリック前に現在のURLセットを記録
        current_s2430_urls = _get_current_s2430_urls(driver)

        # ★修正②：A/button限定で >> ボタンを探す
        next_btn = _find_next_button(driver, debug)

        if next_btn is None:
            if debug:
                st.info(f"ページ{page_num}：「>>」ボタンが見つかりません → 収集終了")
            break

        # クリック実行
        driver.execute_script("arguments[0].click();", next_btn)

        # ★修正③：固定sleepではなくDOM変化を検知して待機
        changed = _wait_for_page_change(driver, current_s2430_urls, timeout=12)

        if not changed:
            if debug:
                st.warning(f"ページ{page_num}：クリック後12秒待ってもページが変化しませんでした → 終了")
                st.image(driver.get_screenshot_as_png(), caption=f"変化なし（ページ{page_num}）")
            break

        page_num += 1

        if debug:
            st.image(driver.get_screenshot_as_png(), caption=f"ページ{page_num} 遷移後")

    return all_urls

# ============================================================
# Step3: 詳細ページからデータ取得
# ============================================================
def extract_val(text, pattern):
    m = re.search(pattern, text, re.DOTALL)
    return m.group(1).strip() if m else ""

def extract_pref(address):
    m = re.search(r"(北海道|東京都|大阪府|京都府|[^\s]{2,3}県)", address)
    return m.group(1) if m else ""

def fetch_detail(driver, name, url):
    store = {"施設名": name, "住所": "", "都道府県": "", "地方": "",
             "薬剤師_常勤": "", "薬剤師_非常勤": "", "総取扱処方箋数": "",
             "詳細URL": url}
    try:
        driver.get(url)
        time.sleep(5)

        txt = driver.find_element(By.TAG_NAME, "body").text
        addr = ""
        for pattern in [
            r"所在地\s*[：:]\s*(.+?)[\n\r]",
            r"所在地\s*\n(.+?)[\n\r]",
            r"〒\d{3}[-－]\d{4}\s*(.+?)[\n\r]",
            r"((?:北海道|東京都|大阪府|京都府|.{2,3}県).{5,50})[\n\r]",
        ]:
            addr = extract_val(txt, pattern)
            if addr and len(addr) > 5:
                break
        store["住所"] = addr.replace("Googleマップで見る","").replace("Google マップで見る","").strip()

        for xpath in ["//*[contains(text(),'実績、結果')]",
                      "//*[contains(text(),'実績・結果')]",
                      "//*[contains(text(),'実績')]"]:
            try:
                tabs = [t for t in driver.find_elements(By.XPATH, xpath) if t.is_displayed()]
                if tabs:
                    driver.execute_script("arguments[0].scrollIntoView(true); arguments[0].click();", tabs[0])
                    time.sleep(3)
                    break
            except Exception:
                pass

        txt = driver.find_element(By.TAG_NAME, "body").text
        store["薬剤師_常勤"]    = extract_val(txt, r"常勤の人数[^\d]*([\d,]+)")
        store["薬剤師_非常勤"]  = extract_val(txt, r"非常勤の人数[^）]*[）)][^\d]*([\d,]+)")
        store["総取扱処方箋数"] = extract_val(txt, r"総取扱処方箋数[^\d]*([\d,]+)")
        if not store["住所"]:
            store["住所"] = extract_val(txt, r"所在地\s*[：:]\s*(.+?)[\n\r]")

        pref = extract_pref(store["住所"])
        store["都道府県"] = pref
        store["地方"]     = REGION_MAP.get(pref, "99_不明")
    except Exception as e:
        store["エラー"] = str(e)
    return store

# ============================================================
# メイン処理
# ============================================================
def run(company, keyword, progress_bar, status_text, debug):
    driver = get_driver()
    all_stores = []
    try:
        status_text.text("🔍 ナビィに接続中...")
        session_id = get_session_id(driver, keyword, debug)
        if not session_id:
            return []
        if debug:
            st.info(f"セッションID取得成功: {session_id}")

        status_text.text("📋 検索結果を収集中...")
        urls = collect_urls(driver, session_id, status_text, debug)
        if not urls:
            st.warning("詳細ページが見つかりませんでした。検索ワードを確認してください。")
            return []

        for i, (name, url) in enumerate(urls):
            status_text.text(f"📊 詳細取得中 {i+1}/{len(urls)}: {name}")
            progress_bar.progress((i+1)/len(urls))
            store = fetch_detail(driver, name, url)
            all_stores.append(store)
            time.sleep(DETAIL_DELAY)
    finally:
        driver.quit()
    return all_stores

def to_excel(stores):
    df = pd.DataFrame(stores)
    sort_cols = [c for c in ["地方","都道府県","施設名"] if c in df.columns]
    df = df.sort_values(sort_cols).reset_index(drop=True)
    col_order = ["施設名","都道府県","地方","住所","薬剤師_常勤","薬剤師_非常勤","総取扱処方箋数","詳細URL"]
    df = df[[c for c in col_order if c in df.columns]]
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="店舗一覧")
        ws = writer.sheets["店舗一覧"]
        from openpyxl.styles import PatternFill, Font, Alignment
        fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w+2, 60)
    buf.seek(0)
    return buf, df

# ============================================================
# Streamlit UI
# ============================================================
st.set_page_config(page_title="薬局情報収集ツール", page_icon="💊", layout="centered")
st.title("💊 薬局情報収集ツール")
st.caption("医療情報ネット（ナビイ）から薬局の薬剤師数・処方箋数を自動収集してExcelに出力します")
st.divider()

col1, col2 = st.columns(2)
with col1:
    company = st.text_input("① 企業名", placeholder="例：株式会社裕生堂")
with col2:
    keyword = st.text_input("② ナビィ検索ワード", placeholder="例：裕生堂薬局")

st.caption("💡 ②はナビィで表示されている店舗ブランド名を入力してください。")

with st.expander("🔧 開発者向けオプション"):
    debug_mode = st.checkbox("デバッグモード", value=False)

st.divider()

if st.button("🔍 収集開始", type="primary", disabled=not (company and keyword), use_container_width=True):
    progress = st.progress(0)
    status   = st.empty()
    with st.spinner(f"「{keyword}」を収集中です。しばらくお待ちください..."):
        try:
            stores = run(company, keyword, progress, status, debug_mode)
        except Exception as e:
            st.error(f"エラーが発生しました: {e}")
            stores = []

    if stores:
        progress.progress(1.0)
        status.empty()
        buf, df = to_excel(stores)
        fname = f"{company}_薬局一覧_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.success(f"✅  {len(stores)} 件を取得しました")
        st.download_button(
            label="📥 Excelをダウンロード",
            data=buf, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True,
        )
        show_cols = [c for c in ["施設名","都道府県","住所","薬剤師_常勤","薬剤師_非常勤","総取扱処方箋数"] if c in df.columns]
        st.subheader("プレビュー（地域別ソート）")
        st.dataframe(df[show_cols], use_container_width=True, hide_index=True)
    else:
        st.warning("結果が取得できませんでした。② の検索ワードを確認してください。")

st.divider()
st.caption("データ出典：医療情報ネット（ナビイ）/ 厚生労働省・都道府県")
