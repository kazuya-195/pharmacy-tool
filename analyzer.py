import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="薬局 生産性分析ツール", page_icon="📊", layout="wide")
st.title("📊 薬局 生産性分析ツール")
st.caption("薬局情報収集ツールで出力したExcelをアップロードして、処方箋枚数・人時生産性を算出・可視化します")
st.divider()

# ============================================================
# ① ファイルアップロード
# ============================================================
uploaded = st.file_uploader("① 収集ツールのExcelをアップロード", type=["xlsx"])
if not uploaded:
    st.info("💡 薬局情報収集ツールでダウンロードしたExcelファイルをアップロードしてください。")
    st.stop()

try:
    df_raw = pd.read_excel(uploaded)
except Exception as e:
    st.error(f"Excelの読み込みに失敗しました: {e}")
    st.stop()

required = ["施設名", "総取扱処方箋数", "薬剤師_常勤"]
missing = [c for c in required if c not in df_raw.columns]
if missing:
    st.error(f"必要な列が見つかりません: {missing}")
    st.stop()

st.success(f"✅ {len(df_raw)} 件読み込みました")
st.divider()

# ============================================================
# ② 営業条件の入力
# ============================================================
st.subheader("② 営業条件を入力（全店舗共通）")
col1, col2 = st.columns(2)
with col1:
    work_days = st.number_input("月間営業日数（日）", min_value=1, max_value=31, value=25, step=1)
with col2:
    work_hours = st.number_input("1日の営業時間（時間）", min_value=0.5, max_value=24.0, value=8.0, step=0.5)
st.divider()

# ============================================================
# ③ 計算
# ============================================================
def to_num(series):
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce"
    )

df = df_raw.copy()
df["総取扱処方箋数_num"] = to_num(df["総取扱処方箋数"])
df["薬剤師_常勤_num"]   = to_num(df["薬剤師_常勤"])
df["月間処方箋枚数"]        = (df["総取扱処方箋数_num"] / 12).round(1)
df["1日処方箋枚数"]         = (df["月間処方箋枚数"] / work_days).round(1)
df["1時間あたり処方箋枚数"] = (df["1日処方箋枚数"] / work_hours).round(1)
df["人時生産性"]            = (df["1時間あたり処方箋枚数"] / df["薬剤師_常勤_num"]).round(2)
df["薬剤師充足率"]          = (df["薬剤師_常勤_num"] / (df["1日処方箋枚数"] / 40)).round(2)
# ※薬剤師充足率：1日40枚/人を適正水準として充足率を算出（1.0=適正、<1.0=不足）

# 欠損警告
no_rx = df["総取扱処方箋数_num"].isna().sum()
no_ph = df["薬剤師_常勤_num"].isna().sum()
if no_rx > 0:
    st.warning(f"⚠️ 総取扱処方箋数が取得できていない店舗: {no_rx} 件")
if no_ph > 0:
    st.warning(f"⚠️ 常勤薬剤師数が取得できていない店舗: {no_ph} 件")

# ============================================================
# ④ 分析結果テーブル
# ============================================================
st.subheader("③ 分析結果一覧")
display_cols = [c for c in ["施設名","都道府県","地方"] if c in df.columns] + [
    "総取扱処方箋数_num","月間処方箋枚数","1日処方箋枚数",
    "1時間あたり処方箋枚数","薬剤師_常勤_num","人時生産性","薬剤師充足率"
]
df_view = df[display_cols].rename(columns={
    "総取扱処方箋数_num":"年間処方箋数","薬剤師_常勤_num":"常勤薬剤師数"
})
st.dataframe(df_view, use_container_width=True, hide_index=True)

# サマリー
valid    = df.dropna(subset=["1日処方箋枚数"])
valid_ph = df.dropna(subset=["人時生産性"])
c1,c2,c3,c4 = st.columns(4)
c1.metric("月間処方箋（平均）",  f"{valid['月間処方箋枚数'].mean():.0f} 枚")
c2.metric("1日処方箋（平均）",   f"{valid['1日処方箋枚数'].mean():.1f} 枚")
c3.metric("1時間あたり（平均）", f"{valid['1時間あたり処方箋枚数'].mean():.1f} 枚")
v = valid_ph["人時生産性"].mean()
c4.metric("人時生産性（平均）",  f"{v:.2f} 枚/人時" if not pd.isna(v) else "—")

st.divider()

# ============================================================
# ⑤ 散布図
# ============================================================
st.subheader("④ 散布図")

df_plot = df.dropna(subset=["1日処方箋枚数","人時生産性"]).copy()
store_names = sorted(df_plot["施設名"].dropna().tolist())

AXIS_OPTIONS = {
    "1日処方箋枚数（枚）":         "1日処方箋枚数",
    "月間処方箋枚数（枚）":        "月間処方箋枚数",
    "1時間あたり処方箋枚数（枚）": "1時間あたり処方箋枚数",
    "人時生産性（枚/人時）":       "人時生産性",
    "常勤薬剤師数（人）":          "薬剤師_常勤_num",
    "薬剤師充足率":                "薬剤師充足率",
}

# --- 設定パネル ---
with st.expander("⚙️ グラフ設定", expanded=True):
    row1 = st.columns(3)
    with row1[0]:
        x_label = st.selectbox("X軸", list(AXIS_OPTIONS.keys()),
                               index=list(AXIS_OPTIONS.keys()).index("1日処方箋枚数（枚）"))
        y_label = st.selectbox("Y軸", list(AXIS_OPTIONS.keys()),
                               index=list(AXIS_OPTIONS.keys()).index("人時生産性（枚/人時）"))
    with row1[1]:
        area_options = ["全店舗"]
        for col in ["地方","都道府県"]:
            if col in df_plot.columns:
                for v in sorted(df_plot[col].dropna().unique()):
                    area_options.append(f"【{col}】{v}")
        area_sel = st.selectbox("エリアで絞り込む", area_options)
        show_zones = st.checkbox("ゾーン（高負荷/余力）を表示", value=True)
    with row1[2]:
        rura_stores = st.multiselect(
            "🌟 RURA設置予定店舗を選択",
            options=store_names,
            placeholder="店舗名を選択..."
        )

x_col = AXIS_OPTIONS[x_label]
y_col = AXIS_OPTIONS[y_label]

# エリア絞り込み
if area_sel == "全店舗":
    df_chart = df_plot.copy()
else:
    col_name = area_sel.split("】")[0].replace("【","")
    val_name = area_sel.split("】")[1]
    df_chart = df_plot[df_plot[col_name] == val_name].copy()

if df_chart.empty:
    st.warning("選択したエリアにデータがありません。")
    st.stop()

df_chart = df_chart.dropna(subset=[x_col, y_col])

# しきい値スライダー
if show_zones:
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        x_thresh = st.slider("X軸しきい値",
            min_value=0.0, max_value=float(df_chart[x_col].max())*1.1,
            value=round(float(df_chart[x_col].median()), 1), step=0.5)
    with col_t2:
        y_thresh = st.slider("Y軸しきい値",
            min_value=0.0, max_value=float(df_chart[y_col].max())*1.1,
            value=round(float(df_chart[y_col].median()), 2), step=0.05)
else:
    x_thresh = float(df_chart[x_col].median())
    y_thresh = float(df_chart[y_col].median())

# ゾーン判定
def zone_label(row):
    hi_x = row[x_col] >= x_thresh
    hi_y = row[y_col] >= y_thresh
    if hi_x and hi_y:     return "高負荷・ターゲット候補"
    if hi_x and not hi_y: return "リソース余力ポテンシャル"
    if not hi_x and hi_y: return "低負荷・高生産性"
    return "低負荷・低生産性"

df_chart["ゾーン"] = df_chart.apply(zone_label, axis=1)
df_chart["RURA予定"] = df_chart["施設名"].isin(rura_stores)

ZONE_COLOR = {
    "高負荷・ターゲット候補":  "#C8860A",
    "リソース余力ポテンシャル": "#1a7a6e",
    "低負荷・高生産性":        "#1F4E79",
    "低負荷・低生産性":        "#999999",
}

# --- Plotly図を構築 ---
fig = go.Figure()

x_max = float(df_chart[x_col].max()) * 1.25
y_max = float(df_chart[y_col].max()) * 1.25

# ゾーン背景
if show_zones:
    fig.add_shape(type="rect", x0=x_thresh, x1=x_max, y0=y_thresh, y1=y_max,
                  fillcolor="#C8860A", opacity=0.08, line_width=0, layer="below")
    fig.add_shape(type="rect", x0=x_thresh, x1=x_max, y0=0, y1=y_thresh,
                  fillcolor="#1a7a6e", opacity=0.08, line_width=0, layer="below")
    fig.add_annotation(x=(x_thresh+x_max)/2, y=y_max*0.95,
                       text="<b>高負荷・ターゲット候補</b>",
                       font=dict(size=12, color="#C8860A"), showarrow=False)
    fig.add_annotation(x=(x_thresh+x_max)/2, y=y_thresh*0.12,
                       text="<b>リソース余力ポテンシャル店舗</b>",
                       font=dict(size=12, color="#1a7a6e"), showarrow=False)
    fig.add_shape(type="line", x0=x_thresh, x1=x_thresh, y0=0, y1=y_max,
                  line=dict(color="#bbb", width=1, dash="dash"))
    fig.add_shape(type="line", x0=0, x1=x_max, y0=y_thresh, y1=y_thresh,
                  line=dict(color="#bbb", width=1, dash="dash"))

# 通常店舗（RURA予定でない）
for zone, group in df_chart[~df_chart["RURA予定"]].groupby("ゾーン"):
    color = ZONE_COLOR.get(zone, "#999")
    hover_parts = []
    for hc in [c for c in ["施設名","都道府県","地方"] if c in group.columns]:
        i = [c for c in ["施設名","都道府県","地方"] if c in group.columns].index(hc)
        hover_parts.append(f"{hc}: %{{customdata[{i}]}}")
    customdata = group[[c for c in ["施設名","都道府県","地方"] if c in group.columns]].values
    hover_tmpl = "<br>".join(hover_parts) + f"<br>{x_label}: %{{x:.1f}}<br>{y_label}: %{{y:.2f}}<extra></extra>"
    fig.add_trace(go.Scatter(
        x=group[x_col], y=group[y_col],
        mode="markers+text",
        name=zone,
        marker=dict(size=10, color=color, line=dict(width=1, color="white")),
        text=group["施設名"],
        textposition="top center",
        textfont=dict(size=9, color=color),
        customdata=customdata,
        hovertemplate=hover_tmpl,
    ))

# RURA設置予定店舗（★で上書き）
if rura_stores:
    rura_group = df_chart[df_chart["RURA予定"]]
    hover_parts = []
    for hc in [c for c in ["施設名","都道府県","地方"] if c in rura_group.columns]:
        i = [c for c in ["施設名","都道府県","地方"] if c in rura_group.columns].index(hc)
        hover_parts.append(f"{hc}: %{{customdata[{i}]}}")
    customdata_r = rura_group[[c for c in ["施設名","都道府県","地方"] if c in rura_group.columns]].values
    hover_tmpl_r = "<br>".join(hover_parts) + f"<br>{x_label}: %{{x:.1f}}<br>{y_label}: %{{y:.2f}}<br>🌟 RURA設置予定<extra></extra>"
    fig.add_trace(go.Scatter(
        x=rura_group[x_col], y=rura_group[y_col],
        mode="markers+text",
        name="🌟 RURA設置予定",
        marker=dict(size=16, symbol="star", color="#E63946",
                    line=dict(width=1.5, color="white")),
        text=rura_group["施設名"],
        textposition="top center",
        textfont=dict(size=10, color="#E63946", family="bold"),
        customdata=customdata_r,
        hovertemplate=hover_tmpl_r,
    ))

fig.update_layout(
    xaxis=dict(title=x_label, range=[0, x_max], gridcolor="#eee"),
    yaxis=dict(title=y_label, range=[0, y_max], gridcolor="#eee"),
    plot_bgcolor="white", paper_bgcolor="white",
    legend=dict(orientation="h", yanchor="bottom", y=-0.28, xanchor="center", x=0.5),
    margin=dict(l=60, r=30, t=30, b=90),
    height=580,
    font=dict(family="Noto Sans JP, sans-serif"),
)

st.plotly_chart(fig, use_container_width=True,
                config={"toImageButtonOptions": {
                    "format": "png", "filename": "pharmacy_scatter",
                    "width": 1400, "height": 800, "scale": 2},
                    "displaylogo": False})
st.caption("💡 グラフ右上の📷アイコンから高解像度PNG保存できます。点にカーソルで店舗詳細が表示されます。")

st.divider()

# ============================================================
# ⑥ 人時生産性ランキング
# ============================================================
st.subheader("⑤ 人時生産性ランキング")
df_rank = df.dropna(subset=["人時生産性"]).copy()
df_rank = df_rank.sort_values("人時生産性", ascending=False).reset_index(drop=True)
df_rank.index += 1

top_n = st.slider("表示件数（上位・下位それぞれ）", min_value=3, max_value=20, value=5)

rank_cols = [c for c in ["施設名","都道府県","地方"] if c in df_rank.columns] + \
            ["1日処方箋枚数","薬剤師_常勤_num","人時生産性"]

col_l, col_r = st.columns(2)
with col_l:
    st.markdown(f"**🔴 人時生産性 TOP {top_n}（高負荷・要注目）**")
    df_top = df_rank.head(top_n)[rank_cols].rename(columns={"薬剤師_常勤_num":"常勤薬剤師数"})
    st.dataframe(df_top, use_container_width=True)
with col_r:
    st.markdown(f"**🔵 人時生産性 BOTTOM {top_n}（余力あり・横展開候補）**")
    df_bot = df_rank.tail(top_n)[rank_cols].rename(columns={"薬剤師_常勤_num":"常勤薬剤師数"})
    st.dataframe(df_bot, use_container_width=True)

st.divider()

# ============================================================
# ⑦ 薬剤師充足率分析
# ============================================================
st.subheader("⑥ 薬剤師充足率分析")
st.caption("1日40枚/人を適正水準として算出。充足率 < 1.0 = 薬剤師が不足している状態")

df_suf = df.dropna(subset=["薬剤師充足率","1日処方箋枚数"]).copy()

# 充足率の分布（棒グラフ）
bins = [0, 0.5, 0.75, 1.0, 1.25, 1.5, float("inf")]
labels = ["〜0.5","0.5〜0.75","0.75〜1.0","1.0〜1.25","1.25〜1.5","1.5〜"]
df_suf["充足率帯"] = pd.cut(df_suf["薬剤師充足率"], bins=bins, labels=labels, right=False)
dist = df_suf["充足率帯"].value_counts().reindex(labels, fill_value=0)

fig2 = go.Figure(go.Bar(
    x=dist.index.tolist(),
    y=dist.values.tolist(),
    marker_color=["#E63946","#E63946","#f4a261","#2a9d8f","#2a9d8f","#1F4E79"],
    text=dist.values.tolist(),
    textposition="outside",
))
fig2.add_vline(x=2.5, line_dash="dash", line_color="#aaa",
               annotation_text="← 不足  適正以上 →", annotation_position="top")
fig2.update_layout(
    xaxis_title="薬剤師充足率（帯域）",
    yaxis_title="店舗数",
    plot_bgcolor="white", paper_bgcolor="white",
    height=320, margin=dict(l=40, r=20, t=20, b=40),
    showlegend=False,
)
st.plotly_chart(fig2, use_container_width=True, config={"displaylogo": False})

# 不足店舗リスト
shortage = df_suf[df_suf["薬剤師充足率"] < 1.0].sort_values("薬剤師充足率")
if not shortage.empty:
    with st.expander(f"⚠️ 薬剤師不足店舗 {len(shortage)} 件を表示"):
        show_cols2 = [c for c in ["施設名","都道府県","地方"] if c in shortage.columns] + \
                     ["1日処方箋枚数","薬剤師_常勤_num","薬剤師充足率"]
        st.dataframe(shortage[show_cols2].rename(columns={"薬剤師_常勤_num":"常勤薬剤師数"}),
                     use_container_width=True, hide_index=True)

st.divider()

# ============================================================
# ⑧ Excel ダウンロード
# ============================================================
out_cols = [c for c in ["施設名","都道府県","地方","住所"] if c in df.columns] + [
    "総取扱処方箋数_num","月間処方箋枚数","1日処方箋枚数",
    "1時間あたり処方箋枚数","薬剤師_常勤_num","人時生産性","薬剤師充足率"
]
df_out = df[out_cols].rename(columns={
    "総取扱処方箋数_num":"年間処方箋数","薬剤師_常勤_num":"常勤薬剤師数"
})
if rura_stores:
    df_out["RURA設置予定"] = df_out["施設名"].isin(rura_stores).map({True:"★予定", False:""})

buf = BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="生産性分析")
    ws = writer.sheets["生産性分析"]
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w+2, 40)
buf.seek(0)

st.download_button(
    label="📥 分析結果をExcelでダウンロード（RURA予定フラグ付き）",
    data=buf,
    file_name=f"薬局生産性分析_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary", use_container_width=True,
)
st.divider()
st.caption(f"計算条件：月間営業日数 {work_days}日 ／ 1日営業時間 {work_hours}時間 ／ 薬剤師適正水準 40枚/人/日")
